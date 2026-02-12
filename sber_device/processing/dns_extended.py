from pathlib import Path
from dataclasses import dataclass, InitVar, field
from typing import ClassVar

import pandas as pd
from openpyxl import load_workbook

from config.configurations import RetailerConfig

dns_retailer_config_extended = RetailerConfig(
    company_name="ДНС",
    table_display_name="DNS_data_extended",
    excel_table_style_name="TableStyleLight9",
    start_summation_row="F",
    table_header_style="Headline 1",
)


@dataclass
class ExtendedReport:
    """
    Обработчик расширенного отчёта ДНС.

    Структура определяется динамически по объединённым ячейкам:
    - «Изделие» — ширина определяет количество полей товара
    - «Итого» — ширина = кол-во метрик, высота = глубина заголовка
    - Магазины — объединённые ячейки переменной ширины с метриками под ними
    """

    df_path: InitVar[Path | str]
    df: pd.DataFrame | None = field(init=False, default=None)
    raw_df: pd.DataFrame | None = field(init=False, default=None)

    # Структура файла (заполняется в _parse_structure)
    _product_fields: list = field(init=False, default_factory=list)
    _shops: list = field(init=False, default_factory=list)
    _data_start_row: int = field(init=False, default=0)

    # Маппинг полей товара из файла → имена выходных колонок
    PRODUCT_FIELD_MAP: ClassVar[dict[str, str]] = {
        "Код": "Артикул",
        "Товар": "Наименование",
        "КодПроизводителя": "Код модели",
    }

    # Маппинг метрик на выходные названия
    METRIC_MAPPING: ClassVar[dict[str, str]] = {
        "Кол-во": "Продажи, шт",
        "Ост. кол-во": "Остатки, шт",
        "Ост. пути": "Остатки в пути",
    }

    # Целочисленные поля
    INTEGER_FIELDS: ClassVar[set[str]] = {
        "Продажи, шт",
        "Остатки, шт",
        "Остатки в пути",
    }

    # Базовые колонки (всегда присутствуют)
    BASE_COLNAMES: ClassVar[list[str]] = [
        "Код модели",
        "Наименование",
        "Артикул",
        "Магазин",
        "Код магазина",
    ]

    # Порядок метрик в выходном файле (максимальный набор)
    METRIC_ORDER: ClassVar[list[str]] = [
        "Продажи, шт",
        "Остатки, шт",
        "Остатки в пути",
        "Себестоимость без НДС",
        "Ост. Сумма без НДС",
        "Продажа без НДС",
    ]

    def __post_init__(self, df_path: Path | str) -> None:
        self._parse_structure(df_path)
        self.raw_df = pd.read_excel(df_path, header=None, dtype=object)
        if self.raw_df is None or self.raw_df.empty:
            raise ValueError("Нет данных для обработки")

    def _parse_structure(self, df_path: Path | str) -> None:
        """
        Определить структуру файла по объединённым ячейкам (openpyxl, 1-based).
        Результаты сохраняются с индексами pandas (0-based).
        """
        wb = load_workbook(df_path)
        ws = wb.active

        # Карта объединений: (row, col) → MergedCellRange
        merge_map = {}
        for m in ws.merged_cells.ranges:
            merge_map[(m.min_row, m.min_col)] = m

        def mwidth(row, col):
            m = merge_map.get((row, col))
            return m.max_col - m.min_col + 1 if m else 1

        def mheight(row, col):
            m = merge_map.get((row, col))
            return m.max_row - m.min_row + 1 if m else 1

        # 1. «Изделие» — первая ячейка R1
        izd_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(1, c).value == "Изделие":
                izd_col = c
                break
        if izd_col is None:
            raise KeyError("Не найден раздел 'Изделие' в первой строке")

        izd_w = mwidth(1, izd_col)
        izd_end = izd_col + izd_w - 1

        # 2. Поля товара — строка под «Изделие», внутри его ширины
        self._product_fields = []
        c = izd_col
        while c <= izd_end:
            val = ws.cell(2, c).value
            if val:
                w = mwidth(2, c)
                self._product_fields.append({
                    "name": str(val),
                    "pd_col": c - 1,  # pandas 0-based
                })
                c += w
            else:
                c += 1

        # 3. «Итого» — сразу после блока «Изделие»
        itogo_col = izd_end + 1
        if ws.cell(1, itogo_col).value != "Итого":
            raise KeyError(
                f"Ожидался 'Итого' в колонке {itogo_col}, "
                f"найдено: {ws.cell(1, itogo_col).value}"
            )
        itogo_w = mwidth(1, itogo_col)
        itogo_h = mheight(1, itogo_col)

        # Строка с именами метрик — сразу под вертикальным объединением «Итого»
        metric_row = 1 + itogo_h

        # 4. Начало данных — строка после метрик
        data_row = metric_row + 1
        # Если первая строка данных — итоги по категории (объединена на всю ширину «Изделие»), пропускаем
        m = merge_map.get((data_row, izd_col))
        if m and (m.max_col - m.min_col + 1) == izd_w:
            data_row += 1
        self._data_start_row = data_row - 1  # pandas 0-based

        # 5. Магазины — после блока «Итого», идём по объединениям
        self._shops = []
        c = itogo_col + itogo_w
        max_col = ws.max_column
        while c <= max_col:
            val = ws.cell(1, c).value
            if val:
                w = mwidth(1, c)
                code = ws.cell(2, c).value

                # Метрики внутри блока магазина (в строке metric_row)
                metrics = {}
                mc = c
                while mc < c + w:
                    mval = ws.cell(metric_row, mc).value
                    if mval:
                        mw = mwidth(metric_row, mc)
                        metrics[str(mval)] = mc - 1  # pandas 0-based
                        mc += mw
                    else:
                        mc += 1

                self._shops.append({
                    "name": str(val),
                    "code": str(code) if code else "",
                    "metrics": metrics,
                })
                c += w
            else:
                c += 1

        wb.close()

    def create_report(self) -> None:
        """
        Сборка отчёта.
        """
        if not self._shops:
            raise ValueError("Не найдено ни одного магазина в данных")

        # Маппинг полей товара: выходное_имя → pandas col
        product_col_map = {}
        for pf in self._product_fields:
            output_name = self.PRODUCT_FIELD_MAP.get(pf["name"])
            if output_name:
                product_col_map[output_name] = pf["pd_col"]

        art_col = product_col_map.get("Артикул", 0)
        all_rows = []

        for row_idx in range(self._data_start_row, len(self.raw_df)):
            artcode = self.raw_df.iloc[row_idx, art_col]

            # Пропускаем строки без числового артикула (итоги, доли и т.п.)
            if pd.isna(artcode):
                continue
            if pd.to_numeric(artcode, errors="coerce") is None or pd.isna(
                pd.to_numeric(artcode, errors="coerce")
            ):
                continue

            # Данные товара
            product_data = {
                name: self.raw_df.iloc[row_idx, col]
                for name, col in product_col_map.items()
            }

            for shop in self._shops:
                row_data = {
                    **product_data,
                    "Магазин": shop["name"],
                    "Код магазина": shop["code"],
                }

                for metric_name, col_idx in shop["metrics"].items():
                    value = self.raw_df.iloc[row_idx, col_idx]
                    output_name = self.METRIC_MAPPING.get(metric_name, metric_name)
                    row_data[output_name] = value if pd.notna(value) else 0

                all_rows.append(row_data)

        self.df = pd.DataFrame(all_rows)

        # Преобразование типов
        self.df["Артикул"] = pd.to_numeric(
            self.df["Артикул"], errors="coerce"
        ).astype(int)
        self.df["Код магазина"] = pd.to_numeric(
            self.df["Код магазина"], errors="coerce"
        ).fillna(0).astype(int)

        for col in self.df.columns:
            if col in self.BASE_COLNAMES:
                continue
            if col in self.INTEGER_FIELDS:
                self.df[col] = pd.to_numeric(
                    self.df[col], errors="coerce"
                ).fillna(0).astype(int)
            else:
                self.df[col] = pd.to_numeric(
                    self.df[col], errors="coerce"
                ).fillna(0)

        # Упорядочиваем: базовые колонки + найденные метрики
        found_metrics = [m for m in self.METRIC_ORDER if m in self.df.columns]
        self.df = self.df.reindex(columns=self.BASE_COLNAMES + found_metrics)

        # Сортировка
        self.df = self.df.sort_values(
            by=["Код модели", "Магазин"]
        ).reset_index(drop=True)


def run_dns_extended(path: Path | str) -> pd.DataFrame:
    """
    Обработка расширенного отчёта ДНС.

    Parameters
    ----------
    path : Path | str
        Путь к файлу с отчётом.

    Returns
    -------
    pd.DataFrame
        Отчёт в формате pandas.DataFrame.
    """
    processor = ExtendedReport(path)
    processor.create_report()
    return processor.df
