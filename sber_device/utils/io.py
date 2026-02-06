from collections import namedtuple
from dataclasses import InitVar, dataclass
from pathlib import Path
import string
from typing import ClassVar

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers, cell_style, named_styles, Font, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension

from config.configurations import Config


DEFAULT_SAVE_PATH = Path("saved")

if not DEFAULT_SAVE_PATH.exists():
    DEFAULT_SAVE_PATH.mkdir()


def read_excel(path: Path | str, header=0, **kwargs) -> pd.DataFrame:
    """
    Чтение Excel-файла по указанному пути.
    """
    return pd.read_excel(path, header=header, **kwargs)


@dataclass
class Writer:
    destination_path: ClassVar[Path | str] = DEFAULT_SAVE_PATH

    input_config: InitVar[Config]
    config: namedtuple = None

    def __post_init__(self, input_config: Config):
        if not self.destination_path.exists():
            self.destination_path.mkdir()

        self.config = input_config.parameters

    def columns_best_fit(self, ws) -> None:
        """
        Make all columns best fit
        """
        column_letters = tuple(
            get_column_letter(col_number) for col_number in range(1, ws.max_column + 1)
        )
        for column_letter in column_letters:
            dim = ColumnDimension(
                ws, index=column_letter, bestFit=True, customWidth=True
            )
            ws.column_dimensions[column_letter] = dim

    def export_to_xls(
        self,
        df: pd.DataFrame,
        fname: str = None,
    ) -> None:
        """
        Экспорт данных в таблицу Excel.

        Parameters:
        -----------
        df : pd.DataFrame
            pandas DataFrame для экспорта.
        fname : str
            Имя файла. По умолчанию None.
            Если не задано, будет сформировано автоматически.

        Returns:
        None
        """

        x, y = df.shape
        right_bound = string.ascii_uppercase[y - 1]

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.config.category}"
        ws.sheet_view.showGridLines = False

        # Make header and an empty line
        ws[
            "A1"
        ].value = f"{self.config.category} в {self.config.company_name} - {self.config.report_period}"
        ws["A1"].style = f"{self.config.table_header_style}"
        ws["A2"] = ""

        # Write dataframe to the file
        for row in dataframe_to_rows(df, header=True, index=False):
            ws.append(row)

        # Make a table
        tab = Table(
            displayName=f"{self.config.table_display_name}",
            ref=f"A3:{right_bound}{x + 4}",
            totalsRowShown=True,
        )

        # Style the table
        style = TableStyleInfo(
            name=f"{self.config.excel_table_style_name}",
            showFirstColumn=True,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Format numbers in the table
        for col in ws.iter_cols(min_col=3, max_col=y, min_row=4, max_row=x + 4):
            for cell in col:
                cell.number_format = numbers.builtin_format_code(3)

        summation_row = ws[
            f"{self.config.start_summation_row}{x + 4}" : f"{get_column_letter(y)}{x + 4}"
        ]

        for cell in summation_row[0]:
            cell.value = (
                f"=SUM({cell.column_letter}{4}:{cell.column_letter}{cell.row - 1})"
            )
            cell.font = Font(bold=True)

        ws[f"A{x + 4}"] = "Итого"
        ws[f"A{x + 4}"].font = Font(bold=True)

        self.columns_best_fit(ws)

        if fname is None:
            fname = f"{self.config.category}_{self.config.company_name}{self.config.report_period}.xlsx"
        wb.save(self.destination_path.joinpath(f"{fname}"))
